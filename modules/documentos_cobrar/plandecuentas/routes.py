# ============================================================
# DXT-CONTA - Módulo Plan de Cuentas
# ============================================================

import re
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from flask import Response, jsonify, render_template, request, session
from psycopg2 import errors

from database.db_manager import DatabaseManager
from modules.plandecuentas import plandecuentas_bp
from utils.decorators import login_required, roles_required


ROLES_LECTURA = [9, 10, 11]
ROLES_EDICION = [9, 10]

TIPOS_CUENTA = [
    'ACTIVO',
    'PASIVO',
    'PATRIMONIO',
    'INGRESO',
    'GASTO',
    'COSTO',
    'ORDEN'
]

NATURALEZAS = [
    'DEUDORA',
    'ACREEDORA'
]


def _json_error(message, status=400, extra=None):
    payload = {'success': False, 'message': message}
    if extra:
        payload.update(extra)
    return jsonify(payload), status


def _json_ok(message=None, **kwargs):
    payload = {'success': True}
    if message:
        payload['message'] = message
    payload.update(kwargs)
    return jsonify(payload)


def _limpiar_texto(value):
    return (value or '').strip()


def _parse_bool(value):
    if isinstance(value, bool):
        return value

    if value is None:
        return False

    return str(value).strip().lower() in ('1', 'true', 't', 'yes', 'si', 'sí', 'on')


def _calcular_nivel(codigo):
    return len(codigo.split('.')) if codigo else 0


def _validar_codigo(codigo):
    if not codigo:
        raise ValueError('El código es obligatorio.')

    if len(codigo) > 30:
        raise ValueError('El código no puede exceder 30 caracteres.')

    if not re.fullmatch(r'[0-9]+(?:\.[0-9]+)*', codigo):
        raise ValueError('El código solo puede contener números y puntos, sin espacios ni letras.')

    if '..' in codigo or codigo.startswith('.') or codigo.endswith('.'):
        raise ValueError('El código tiene un formato inválido.')


def _naturaleza_sugerida(tipo):
    mapa = {
        'ACTIVO': 'DEUDORA',
        'GASTO': 'DEUDORA',
        'COSTO': 'DEUDORA',
        'PASIVO': 'ACREEDORA',
        'PATRIMONIO': 'ACREEDORA',
        'INGRESO': 'ACREEDORA',
        'ORDEN': 'DEUDORA'
    }
    return mapa.get(tipo, 'DEUDORA')
    
def _build_filtros_listado(args):
    codigo = _limpiar_texto(args.get('codigo'))
    nombre = _limpiar_texto(args.get('nombre'))
    tipo = _limpiar_texto(args.get('tipo'))
    nivel = _limpiar_texto(args.get('nivel'))
    activo = _limpiar_texto(args.get('activo'))
    postable = _limpiar_texto(args.get('postable'))

    condiciones = []
    params = []

    if codigo:
        condiciones.append('c.codigo ILIKE %s')
        params.append(f'%{codigo}%')

    if nombre:
        condiciones.append('c.nombre ILIKE %s')
        params.append(f'%{nombre}%')

    if tipo:
        condiciones.append('c.tipo = %s')
        params.append(tipo)

    if nivel:
        try:
            nivel_num = int(nivel)
            condiciones.append('c.nivel = %s')
            params.append(nivel_num)
        except ValueError:
            raise ValueError('El filtro nivel es inválido.')

    if activo in ('true', 'false'):
        condiciones.append('c.activo = %s')
        params.append(activo == 'true')

    if postable in ('true', 'false'):
        condiciones.append('c.es_postable = %s')
        params.append(postable == 'true')

    where_sql = f"WHERE {' AND '.join(condiciones)}" if condiciones else ''
    return where_sql, tuple(params)


def _fetch_plan_cuentas_rows(db, where_sql='', params=()):
    return db.execute_query(f"""
        SELECT
            c.codigo,
            c.nombre,
            c.nivel,
            c.tipo,
            c.naturaleza,
            c.es_postable,
            c.requiere_auxiliar,
            c.requiere_cc,
            c.codigo_padre,
            c.activo,
            c.creado_en,
            c.actualizado_en,
            EXISTS (
                SELECT 1
                FROM contabilidad.cuenta h
                WHERE h.codigo_padre = c.codigo
            ) AS tiene_hijos
        FROM contabilidad.cuenta c
        {where_sql}
        ORDER BY c.codigo
    """, params)


def _bool_excel(value):
    return 'Sí' if bool(value) else 'No'

def _obtener_cuenta(db, codigo):
    rows = db.execute_query("""
        SELECT
            codigo,
            nombre,
            nivel,
            tipo,
            naturaleza,
            es_postable,
            requiere_auxiliar,
            requiere_cc,
            codigo_padre,
            activo,
            creado_en,
            actualizado_en
        FROM contabilidad.cuenta
        WHERE codigo = %s
        LIMIT 1
    """, (codigo,))
    return rows[0] if rows else None




def _relation_column_exists(db, schema, relation, column):
    rows = db.execute_query("""
        SELECT 1
        FROM pg_attribute a
        INNER JOIN pg_class c ON c.oid = a.attrelid
        INNER JOIN pg_namespace n ON n.oid = c.relnamespace
        WHERE n.nspname = %s
          AND c.relname = %s
          AND a.attname = %s
          AND a.attnum > 0
          AND NOT a.attisdropped
        LIMIT 1
    """, (schema, relation, column))
    return bool(rows)

def _tiene_hijos(db, codigo):
    rows = db.execute_query("""
        SELECT 1
        FROM contabilidad.cuenta
        WHERE codigo_padre = %s
        LIMIT 1
    """, (codigo,))
    return bool(rows)


def _dependencias_cuenta(db, codigo):
    checks = [
        ('asiento_detalle', 'contabilidad', 'asiento_detalle', 'cuenta_codigo'),
        ('auxiliar_cuenta', 'contabilidad', 'auxiliar_cuenta', 'cuenta_codigo'),
        ('caja', 'contabilidad', 'caja', 'cuenta_contable_codigo'),
        ('cuenta_bancaria', 'contabilidad', 'cuenta_bancaria', 'cuenta_contable_codigo'),
        ('compra_detalle', 'contabilidad', 'compra_detalle', 'cuenta_gasto_codigo'),
        ('compromiso', 'contabilidad', 'compromiso', 'cuenta_contable'),
        ('venta', 'contabilidad', 'venta', 'contra_cuenta_codigo'),
        ('compra', 'contabilidad', 'compra', 'contra_cuenta_codigo'),
        ('cobro', 'contabilidad', 'cobro', 'contra_cuenta_codigo'),
        ('pago', 'contabilidad', 'pago', 'contra_cuenta_codigo'),
        ('movimiento_tesoreria', 'contabilidad', 'movimiento_tesoreria', 'contra_cuenta_codigo'),
        ('venta_detalle', 'contabilidad', 'venta_detalle', 'cuenta_ingreso_codigo'),
    ]

    dependencias = []

    for nombre, schema, relation, columna in checks:
        if not _relation_column_exists(db, schema, relation, columna):
            continue

        tabla = f'{schema}.{relation}'
        sql = f'SELECT COUNT(*) AS total FROM {tabla} WHERE {columna} = %s'
        rows = db.execute_query(sql, (codigo,))
        total = rows[0]['total'] if rows else 0
        if total and int(total) > 0:
            dependencias.append({
                'nombre': nombre,
                'tabla': tabla,
                'total': int(total)
            })

    return dependencias


def _validar_estructura(db, codigo, codigo_padre, tipo, naturaleza, es_postable, codigo_original=None):
    _validar_codigo(codigo)

    nivel = _calcular_nivel(codigo)

    if nivel < 1:
        raise ValueError('El nivel calculado del código es inválido.')

    if tipo not in TIPOS_CUENTA:
        raise ValueError('El tipo de cuenta es inválido.')

    if naturaleza not in NATURALEZAS:
        raise ValueError('La naturaleza es inválida.')

    if codigo_padre:
        padre = _obtener_cuenta(db, codigo_padre)
        if not padre:
            raise ValueError('La cuenta padre no existe.')

        if not padre['activo']:
            raise ValueError('No se puede usar una cuenta padre inactiva.')

        if padre['es_postable']:
            raise ValueError('No se pueden crear hijos debajo de una cuenta postable.')

        if tipo != padre['tipo']:
            raise ValueError('El tipo de la cuenta hija debe coincidir con el tipo de la cuenta padre.')

        if naturaleza != padre['naturaleza']:
            raise ValueError('La naturaleza de la cuenta hija debe coincidir con la naturaleza de la cuenta padre.')

        if not codigo.startswith(f'{codigo_padre}.'):
            raise ValueError('El código debe comenzar con el código de la cuenta padre seguido de un punto.')

        if nivel != padre['nivel'] + 1:
            raise ValueError('El nivel de la cuenta no coincide con el nivel esperado según la cuenta padre.')

        resto = codigo[len(codigo_padre) + 1:]
        if '.' in resto or not resto:
            raise ValueError('El código hijo debe agregar un único segmento respecto al padre.')

    else:
        if nivel != 1:
            raise ValueError('Las cuentas sin padre solo pueden ser de nivel 1.')

    if codigo_original and codigo_original != codigo:
        raise ValueError('No está permitido cambiar el código de una cuenta existente en esta versión.')

    if es_postable and _tiene_hijos(db, codigo_original or codigo):
        raise ValueError('Una cuenta con hijos no puede marcarse como postable.')

    return nivel


@plandecuentas_bp.route('/')
@login_required
@roles_required(ROLES_LECTURA)
def index():
    """Pantalla principal del Plan de Cuentas."""
    puede_editar = int(session.get('rol_id') or 0) in ROLES_EDICION

    return render_template(
        'plandecuentas_index.html',
        tipos_cuenta=TIPOS_CUENTA,
        naturalezas=NATURALEZAS,
        puede_editar=puede_editar
    )


@plandecuentas_bp.route('/api/listar', methods=['GET'])
@login_required
@roles_required(ROLES_LECTURA)
def listar():
    """Lista cuentas del catálogo con filtros."""
    try:
        where_sql, params = _build_filtros_listado(request.args)

        with DatabaseManager() as db:
            rows = _fetch_plan_cuentas_rows(db, where_sql, params)

        data = []
        for row in rows:
            data.append({
                'codigo': row['codigo'],
                'nombre': row['nombre'],
                'nivel': row['nivel'],
                'tipo': row['tipo'],
                'naturaleza': row['naturaleza'],
                'es_postable': row['es_postable'],
                'requiere_auxiliar': row['requiere_auxiliar'],
                'requiere_cc': row['requiere_cc'],
                'codigo_padre': row['codigo_padre'],
                'activo': row['activo'],
                'tiene_hijos': row['tiene_hijos'],
            })

        return jsonify({'data': data})

    except ValueError as e:
        return _json_error(str(e))
    except Exception as e:
        return _json_error(f'No se pudo listar el plan de cuentas: {str(e)}', 500)

@plandecuentas_bp.route('/api/exportar-excel', methods=['GET'])
@login_required
@roles_required(ROLES_LECTURA)
def exportar_excel():
    try:
        where_sql, params = _build_filtros_listado(request.args)

        with DatabaseManager() as db:
            rows = _fetch_plan_cuentas_rows(db, where_sql, params)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Plan de Cuentas'

        titulo_fill = PatternFill(fill_type='solid', fgColor='0F2340')
        header_fill = PatternFill(fill_type='solid', fgColor='DCEAF9')
        white_font = Font(color='FFFFFF', bold=True, size=14)
        header_font = Font(bold=True, size=10)
        normal_font = Font(size=10)
        center = Alignment(horizontal='center', vertical='center')
        left = Alignment(horizontal='left', vertical='center')
        right = Alignment(horizontal='right', vertical='center')
        thin = Side(style='thin', color='D9E2EC')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        filtros_resumen = []
        if _limpiar_texto(request.args.get('codigo')):
            filtros_resumen.append(f'Código: {_limpiar_texto(request.args.get("codigo"))}')
        if _limpiar_texto(request.args.get('nombre')):
            filtros_resumen.append(f'Nombre: {_limpiar_texto(request.args.get("nombre"))}')
        if _limpiar_texto(request.args.get('tipo')):
            filtros_resumen.append(f'Tipo: {_limpiar_texto(request.args.get("tipo"))}')
        if _limpiar_texto(request.args.get('nivel')):
            filtros_resumen.append(f'Nivel: {_limpiar_texto(request.args.get("nivel"))}')
        if _limpiar_texto(request.args.get('activo')):
            filtros_resumen.append(
                'Activo: Sí' if _limpiar_texto(request.args.get('activo')) == 'true' else 'Activo: No'
            )
        if _limpiar_texto(request.args.get('postable')):
            filtros_resumen.append(
                'Postable: Sí' if _limpiar_texto(request.args.get('postable')) == 'true' else 'Postable: No'
            )

        ws.merge_cells('A1:J1')
        ws['A1'] = 'PLAN DE CUENTAS'
        ws['A1'].fill = titulo_fill
        ws['A1'].font = white_font
        ws['A1'].alignment = center

        ws.merge_cells('A2:J2')
        ws['A2'] = f'Emitido: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        ws['A2'].alignment = center
        ws['A2'].font = Font(italic=True, size=10)

        ws.merge_cells('A3:J3')
        ws['A3'] = 'Filtros: ' + (' | '.join(filtros_resumen) if filtros_resumen else 'Sin filtros')
        ws['A3'].alignment = left
        ws['A3'].font = Font(size=10)

        headers = [
            'Código',
            'Nombre',
            'Nivel',
            'Tipo',
            'Naturaleza',
            'Postable',
            'Requiere Auxiliar',
            'Requiere C.C.',
            'Padre',
            'Activo',
        ]

        row_header = 5
        for col_idx, label in enumerate(headers, start=1):
            cell = ws.cell(row=row_header, column=col_idx, value=label)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        current_row = row_header + 1

        for row in rows:
            nombre_indentado = f'{"    " * max(0, int(row["nivel"]) - 1)}{row["nombre"]}'

            values = [
                row['codigo'],
                nombre_indentado,
                row['nivel'],
                row['tipo'],
                row['naturaleza'],
                _bool_excel(row['es_postable']),
                _bool_excel(row['requiere_auxiliar']),
                _bool_excel(row['requiere_cc']),
                row['codigo_padre'] or '',
                _bool_excel(row['activo']),
            ]

            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.font = normal_font
                cell.border = border

                if col_idx in (3, 6, 7, 8, 10):
                    cell.alignment = center
                else:
                    cell.alignment = left

            current_row += 1

        widths = {
            'A': 18,
            'B': 48,
            'C': 10,
            'D': 16,
            'E': 16,
            'F': 12,
            'G': 18,
            'H': 15,
            'I': 18,
            'J': 10,
        }

        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        ws.freeze_panes = 'A6'

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        nombre_archivo = f'plan_de_cuentas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename={nombre_archivo}'
            }
        )

    except ValueError as e:
        return _json_error(str(e))
    except Exception as e:
        return _json_error(f'No se pudo exportar el plan de cuentas a Excel: {str(e)}', 500)

@plandecuentas_bp.route('/api/obtener/<path:codigo>', methods=['GET'])
@login_required
@roles_required(ROLES_LECTURA)
def obtener(codigo):
    try:
        with DatabaseManager() as db:
            cuenta = _obtener_cuenta(db, codigo)
            if not cuenta:
                return _json_error('La cuenta solicitada no existe.', 404)

            cuenta['tiene_hijos'] = _tiene_hijos(db, codigo)
            cuenta['dependencias'] = _dependencias_cuenta(db, codigo)

        return _json_ok(data=cuenta)

    except Exception as e:
        return _json_error(f'No se pudo obtener la cuenta: {str(e)}', 500)


@plandecuentas_bp.route('/api/padres', methods=['GET'])
@login_required
@roles_required(ROLES_LECTURA)
def listar_padres():
    exclude_codigo = _limpiar_texto(request.args.get('exclude_codigo'))

    try:
        with DatabaseManager() as db:
            params = []
            where = ["activo = true", "es_postable = false"]

            if exclude_codigo:
                where.append("codigo <> %s")
                params.append(exclude_codigo)
            rows = db.execute_query(f"""
                SELECT codigo, nombre, nivel, tipo, naturaleza
                FROM contabilidad.cuenta
                WHERE {' AND '.join(where)}
                ORDER BY codigo
            """, tuple(params))

        return _json_ok(data=rows)

    except Exception as e:
        return _json_error(f'No se pudo obtener la lista de cuentas padre: {str(e)}', 500)


@plandecuentas_bp.route('/api/crear', methods=['POST'])
@login_required
@roles_required(ROLES_EDICION)
def crear():
    data = request.get_json() or {}

    codigo = _limpiar_texto(data.get('codigo'))
    nombre = _limpiar_texto(data.get('nombre'))
    tipo = _limpiar_texto(data.get('tipo'))
    naturaleza = _limpiar_texto(data.get('naturaleza'))
    codigo_padre = _limpiar_texto(data.get('codigo_padre')) or None
    es_postable = _parse_bool(data.get('es_postable'))
    requiere_auxiliar = _parse_bool(data.get('requiere_auxiliar'))
    requiere_cc = _parse_bool(data.get('requiere_cc'))
    activo = _parse_bool(data.get('activo', True))

    if not nombre:
        return _json_error('El nombre es obligatorio.')

    if len(nombre) > 250:
        return _json_error('El nombre no puede exceder 250 caracteres.')

    try:
        with DatabaseManager() as db:
            nivel = _validar_estructura(
                db=db,
                codigo=codigo,
                codigo_padre=codigo_padre,
                tipo=tipo,
                naturaleza=naturaleza,
                es_postable=es_postable
            )

            existente = _obtener_cuenta(db, codigo)
            if existente:
                return _json_error('Ya existe una cuenta con ese código.')

            query = """
                INSERT INTO contabilidad.cuenta (
                    codigo,
                    nombre,
                    nivel,
                    tipo,
                    naturaleza,
                    es_postable,
                    requiere_auxiliar,
                    requiere_cc,
                    codigo_padre,
                    activo
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
            params = (
                codigo,
                nombre,
                nivel,
                tipo,
                naturaleza,
                es_postable,
                requiere_auxiliar,
                requiere_cc,
                codigo_padre,
                activo
            )
            db.execute_insert(query, params, return_id=False)

        return _json_ok('Cuenta creada correctamente.')

    except ValueError as e:
        return _json_error(str(e))
    except errors.UniqueViolation:
        return _json_error('Ya existe una cuenta con ese código.')
    except Exception as e:
        return _json_error(f'No se pudo crear la cuenta: {str(e)}', 500)


@plandecuentas_bp.route('/api/actualizar/<path:codigo>', methods=['PUT'])
@login_required
@roles_required(ROLES_EDICION)
def actualizar(codigo):
    data = request.get_json() or {}

    nombre = _limpiar_texto(data.get('nombre'))
    tipo = _limpiar_texto(data.get('tipo'))
    naturaleza = _limpiar_texto(data.get('naturaleza'))
    codigo_padre = _limpiar_texto(data.get('codigo_padre')) or None
    es_postable = _parse_bool(data.get('es_postable'))
    requiere_auxiliar = _parse_bool(data.get('requiere_auxiliar'))
    requiere_cc = _parse_bool(data.get('requiere_cc'))
    activo = _parse_bool(data.get('activo', True))

    if not nombre:
        return _json_error('El nombre es obligatorio.')

    try:
        with DatabaseManager() as db:
            cuenta_actual = _obtener_cuenta(db, codigo)
            if not cuenta_actual:
                return _json_error('La cuenta que intenta actualizar no existe.', 404)

            _validar_estructura(
                db=db,
                codigo=cuenta_actual['codigo'],
                codigo_padre=codigo_padre,
                tipo=tipo,
                naturaleza=naturaleza,
                es_postable=es_postable,
                codigo_original=cuenta_actual['codigo']
            )

            if cuenta_actual['tiene_hijos'] if 'tiene_hijos' in cuenta_actual else _tiene_hijos(db, codigo):
                if es_postable:
                    return _json_error('Una cuenta con hijos no puede ser postable.')

            deps = _dependencias_cuenta(db, codigo)
            if deps:
                if tipo != cuenta_actual['tipo']:
                    return _json_error('No se puede cambiar el tipo de una cuenta que ya tiene uso en otros registros.')
                if naturaleza != cuenta_actual['naturaleza']:
                    return _json_error('No se puede cambiar la naturaleza de una cuenta que ya tiene uso en otros registros.')
                if codigo_padre != cuenta_actual['codigo_padre']:
                    return _json_error('No se puede cambiar la cuenta padre de una cuenta que ya tiene uso en otros registros.')

            query = """
                UPDATE contabilidad.cuenta
                SET
                    nombre = %s,
                    tipo = %s,
                    naturaleza = %s,
                    es_postable = %s,
                    requiere_auxiliar = %s,
                    requiere_cc = %s,
                    codigo_padre = %s,
                    activo = %s,
                    actualizado_en = CURRENT_TIMESTAMP
                WHERE codigo = %s
            """
            params = (
                nombre,
                tipo,
                naturaleza,
                es_postable,
                requiere_auxiliar,
                requiere_cc,
                codigo_padre,
                activo,
                codigo
            )
            db.execute_update(query, params)

        return _json_ok('Cuenta actualizada correctamente.')

    except ValueError as e:
        return _json_error(str(e))
    except Exception as e:
        return _json_error(f'No se pudo actualizar la cuenta: {str(e)}', 500)


@plandecuentas_bp.route('/api/eliminar/<path:codigo>', methods=['DELETE'])
@login_required
@roles_required(ROLES_EDICION)
def eliminar(codigo):
    try:
        with DatabaseManager() as db:
            cuenta = _obtener_cuenta(db, codigo)
            if not cuenta:
                return _json_error('La cuenta no existe.', 404)

            if _tiene_hijos(db, codigo):
                return _json_error(
                    'No se puede eliminar la cuenta porque tiene subcuentas asociadas.',
                    extra={'bloqueo': 'hijos'}
                )

            dependencias = _dependencias_cuenta(db, codigo)
            if dependencias:
                return _json_error(
                    'No se puede eliminar la cuenta porque tiene registros relacionados. Puede desactivarla.',
                    extra={'bloqueo': 'dependencias', 'dependencias': dependencias}
                )

            db.execute_delete(
                "DELETE FROM contabilidad.cuenta WHERE codigo = %s",
                (codigo,)
            )

        return _json_ok('Cuenta eliminada correctamente.')

    except Exception as e:
        return _json_error(f'No se pudo eliminar la cuenta: {str(e)}', 500)
    
# ------------------------------------------------------------
# AYUDA DEL MÓDULO
# ------------------------------------------------------------
@plandecuentas_bp.route('/help')
@login_required
@roles_required(ROLES_LECTURA)
def help():
    return render_template('plandecuentas_help.html')