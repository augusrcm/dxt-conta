# ============================================================
# DXT CONTA - Reportes Rápidos - Exportación Excel genérica
# ============================================================

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from modules.reportes_rapidos.core.utils import decimal_value


def _money_note_display(value):
    note = str(value or '').strip()
    if not note:
        return ''
    if not note.endswith('.'):
        note += '.'
    if note.startswith('(') and note.endswith(')'):
        return note
    return f'({note})'


def build_excel(report, payload):
    wb = Workbook()
    ws = wb.active
    ws.title = getattr(report, 'WORKSHEET_TITLE', payload.get('titulo', 'Reporte'))[:31]

    fill_title = PatternFill('solid', fgColor='0F2340')
    fill_header = PatternFill('solid', fgColor='EAF1F8')
    fill_critical = PatternFill('solid', fgColor='FDE2E2')
    fill_high = PatternFill('solid', fgColor='FFF0D6')
    fill_medium = PatternFill('solid', fgColor='FFF8D6')
    fill_low = PatternFill('solid', fgColor='EAF6EF')
    font_title = Font(color='FFFFFF', bold=True, size=14)
    font_header = Font(color='0F2340', bold=True)
    border = Border(
        left=Side(style='thin', color='D9E1EA'),
        right=Side(style='thin', color='D9E1EA'),
        top=Side(style='thin', color='D9E1EA'),
        bottom=Side(style='thin', color='D9E1EA'),
    )

    columns = report.excel_columns()
    last_col = len(columns)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws.cell(row=1, column=1, value=payload.get('titulo', 'Reporte'))
    ws.cell(row=1, column=1).fill = fill_title
    ws.cell(row=1, column=1).font = font_title
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    ws.cell(
        row=2,
        column=1,
        value=(
            f"{payload.get('descripcion_periodo', '')} · "
            f"Unidad: {payload.get('unidad_label', '')} · "
            f"Emitido: {payload.get('emitido_en', '')}"
        ),
    )
    ws.cell(row=2, column=1).alignment = Alignment(horizontal='center')

    summary = payload.get('summary', {})
    moneda_note = _money_note_display(summary.get('moneda_display_note') or '')
    resumen_texto = report.excel_summary_text(summary)
    if moneda_note:
        resumen_texto = f'{moneda_note} · {resumen_texto}' if resumen_texto else moneda_note
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_col)
    ws.cell(row=3, column=1, value=resumen_texto)
    ws.cell(row=3, column=1).alignment = Alignment(horizontal='center')

    criterio = str(payload.get('criterio_reporte') or '').strip()
    fuente = str(payload.get('fuente_datos') or '').strip()
    criterio_texto = ''
    if criterio:
        criterio_texto = f'Criterio: {criterio}'
        if fuente:
            criterio_texto = f'{criterio_texto} Fuente: {fuente}'

    header_row = 6 if criterio_texto else 5
    if criterio_texto:
        ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=last_col)
        ws.cell(row=4, column=1, value=criterio_texto)
        ws.cell(row=4, column=1).alignment = Alignment(horizontal='center', wrap_text=True)
        ws.cell(row=4, column=1).font = Font(color='5F6F83', italic=True)

    for col_idx, (_, label, width) in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.fill = fill_header
        cell.font = font_header
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    fill_by_priority = {
        'CRITICA': fill_critical,
        'ALTA': fill_high,
        'MEDIA': fill_medium,
        'BAJA': fill_low,
    }

    money_fields = set(getattr(report, 'MONEY_FIELDS', {'monto'}))

    for row_idx, item in enumerate(payload.get('rows', []), start=header_row + 1):
        fill = fill_by_priority.get(item.get('prioridad_codigo'))
        for col_idx, (field, _, _) in enumerate(columns, start=1):
            value = item.get(field)
            is_money = field in money_fields or str(field).startswith('monto')
            if is_money:
                value = float(decimal_value(value))
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            if is_money:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='top')
            if fill:
                cell.fill = fill

    last_row = max(header_row + 1, header_row + len(payload.get('rows', [])))
    ws.freeze_panes = 'A7' if criterio_texto else 'A6'
    ws.auto_filter.ref = f'A{header_row}:{get_column_letter(last_col)}{last_row}'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
